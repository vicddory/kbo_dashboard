"""
OPS+ 계산 모듈
공식: OPS+ = 100 × (OBP/lgOBP + SLG/lgSLG - 1) / (PF/100)
출처: Baseball-Reference, FanGraphs

100 = 리그 평균, 150 = 평균보다 50% 우수
"""

import sqlite3
import os
from openpyxl import load_workbook

TEAM_NAME_MAP = {
    (1985, '청보'): '삼미/청보',
    (2001, 'KIA'): '해태/KIA',
    (2009, '히어로즈'): '우리',
}

PF_2025 = {
    '삼성': 117.4, '롯데': 112.7, 'NC': 111.2,
    'SSG': 107.4, 'KT': 107.0, '한화': 106.5,
    'LG': 93.3, 'KIA': 93.0, '두산': 87.6, '키움': 77.2,
}


def load_park_factors(excel_files):
    pf_all = {}
    for f in excel_files:
        if not os.path.exists(f):
            continue
        wb = load_workbook(f, data_only=True)
        for name in wb.sheetnames:
            if '파크팩터' in name:
                year = int(name.split('_')[1])
                ws = wb[name]
                pf_all[year] = {}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[0] != '팀':
                        pf_all[year][row[0]] = float(row[-1])
        wb.close()
    pf_all[2025] = PF_2025
    return pf_all


def get_pf(pf_data, year, team):
    if year in pf_data and team in pf_data[year]:
        return pf_data[year][team]
    mapped = TEAM_NAME_MAP.get((year, team))
    if mapped and year in pf_data and mapped in pf_data[year]:
        return pf_data[year][mapped]
    return None


def calculate_ops_plus(db_path, pf_data, min_pa_ratio=3.1):
    """
    전 연도 OPS+ 계산

    Args:
        db_path: kbo_data.db 경로
        pf_data: 파크팩터 데이터
        min_pa_ratio: 규정타석 = 팀 경기수 × 이 값 (MLB 기준 3.1)

    Returns:
        {year: [{"Player_ID", "선수명", "팀명", "OBP", "SLG", "OPS", "OPS+", ...}]}
    """
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    results = {}

    for year in range(1982, 2026):
        t1 = f'kbo_hitting_basic1_{year}'
        t2 = f'kbo_hitting_basic2_{year}'

        # 테이블 존재 확인
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (t1,))
        if not cursor.fetchone():
            continue
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (t2,))
        if not cursor.fetchone():
            continue

        # basic1: PA, AB, H, 2B, 3B, HR, TB, R, SF, SAC, G
        cursor.execute(f"""
            SELECT Player_ID, 선수명, 팀명, PA, AB, H, '2B', '3B', HR, TB, R, RBI, SF, SAC, G, AVG
            FROM [{t1}]
        """)
        basic1 = {row[0]: row for row in cursor.fetchall()}

        # basic2: BB, IBB, HBP, SO, SLG, OBP, OPS
        cursor.execute(f"""
            SELECT Player_ID, BB, IBB, HBP, SO, SLG, OBP, OPS
            FROM [{t2}]
        """)
        basic2 = {row[0]: row for row in cursor.fetchall()}

        if not basic1 or not basic2:
            continue

        # 리그 평균 계산 (합산 방식)
        total_h, total_ab, total_bb, total_hbp, total_sf, total_tb = 0, 0, 0, 0, 0, 0
        for pid, row1 in basic1.items():
            if pid not in basic2:
                continue
            row2 = basic2[pid]
            h = row1[5] or 0
            ab = row1[4] or 0
            tb = row1[9] or 0
            sf = row1[12] or 0
            bb = row2[1] or 0
            hbp = row2[3] or 0
            total_h += h
            total_ab += ab
            total_bb += bb
            total_hbp += hbp
            total_sf += sf
            total_tb += tb

        if total_ab == 0:
            continue

        denom_obp = total_ab + total_bb + total_hbp + total_sf
        lg_obp = (total_h + total_bb + total_hbp) / denom_obp if denom_obp > 0 else 0
        lg_slg = total_tb / total_ab if total_ab > 0 else 0

        if lg_obp == 0 or lg_slg == 0:
            continue

        # 규정타석 계산
        max_g = max((row[14] for row in basic1.values() if row[14]), default=144)
        min_pa = max_g * min_pa_ratio

        # 개인 OPS+ 계산
        year_results = []
        for pid, row1 in basic1.items():
            if pid not in basic2:
                continue
            row2 = basic2[pid]

            pa = row1[3] or 0
            if pa < min_pa:
                continue

            team = row1[2]
            obp = row2[6]
            slg = row2[5]
            ops = row2[7]

            if obp is None or slg is None or obp == 0:
                continue

            pf = get_pf(pf_data, year, team)
            if pf is None:
                continue

            # OPS+ = 100 × (OBP/lgOBP + SLG/lgSLG - 1) / (PF/100)
            ops_plus = 100 * (obp / lg_obp + slg / lg_slg - 1) / (pf / 100)

            year_results.append({
                'Player_ID': pid,
                '선수명': row1[1],
                '팀명': team,
                'G': row1[14],
                'PA': pa,
                'AVG': row1[15],
                'OBP': obp,
                'SLG': slg,
                'OPS': ops,
                'HR': row1[8],
                'RBI': row1[11],
                'R': row1[10],
                'lgOBP': round(lg_obp, 3),
                'lgSLG': round(lg_slg, 3),
                'PF': pf,
                'OPS+': round(ops_plus, 1),
            })

        year_results.sort(key=lambda x: -x['OPS+'])
        results[year] = year_results

    conn.close()
    return results


def print_ops_plus(results, year, top_n=20):
    if year not in results or not results[year]:
        print(f"{year}년 데이터 없음")
        return
    data = results[year]
    lg_obp = data[0]['lgOBP']
    lg_slg = data[0]['lgSLG']
    print(f"\n{'='*75}")
    print(f"  {year} KBO OPS+ (리그 OBP: {lg_obp}, 리그 SLG: {lg_slg})")
    print(f"{'='*75}")
    print(f"{'#':>3} {'선수명':>8} {'팀':>6} {'AVG':>6} {'OBP':>6} {'SLG':>6} {'OPS':>6} {'HR':>4} {'PF':>6} {'OPS+':>6}")
    print(f"{'-'*75}")
    for i, p in enumerate(data[:top_n], 1):
        print(f"{i:>3} {p['선수명']:>8} {p['팀명']:>6} {p['AVG']:>6.3f} {p['OBP']:>6.3f} "
              f"{p['SLG']:>6.3f} {p['OPS']:>6.3f} {p['HR']:>4} {p['PF']:>6.1f} {p['OPS+']:>6.1f}")


if __name__ == "__main__":
    pf_files = [
        "kbo_statiz_data_1982_1998_.xlsx",
        "kbo_statiz_data_1999_2009_.xlsx",
        "kbo_statiz_data_2010_2021_.xlsx",
        "kbo_statiz_data_2022_2024_.xlsx",
    ]
    pf_data = load_park_factors(pf_files)
    print(f"파크팩터: {len(pf_data)}개 연도")

    results = calculate_ops_plus("kbo_data.db", pf_data)
    print(f"OPS+ 계산: {len(results)}개 연도")

    for year in [2025, 2024, 2020, 2015, 2010, 2005, 2000, 1995, 1990, 1985]:
        if year in results:
            print_ops_plus(results, year, top_n=10)
