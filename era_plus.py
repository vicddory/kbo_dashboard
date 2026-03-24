"""
ERA+ 계산 모듈
공식: ERA+ = 100 × (lgERA / ERA) × (PF / 100)
출처: Wikipedia Adjusted ERA+ (현재 표준), FanGraphs Sabermetrics Library
"""

import sqlite3
import os
from openpyxl import load_workbook

TEAM_NAME_MAP = {
    (1985, '청보'): '삼미/청보',
    (2001, 'KIA'): '해태/KIA',
    (2009, '히어로즈'): '우리',
}

PF_2025 = {
    '삼성': 117.4, '롯데': 112.7, 'NC': 111.2,
    'SSG': 107.4, 'KT': 107.0, '한화': 106.5,
    'LG': 93.3, 'KIA': 93.0, '두산': 87.6, '키움': 77.2,
}


def load_park_factors(excel_files):
    pf_all = {}
    for f in excel_files:
        if not os.path.exists(f):
            continue
        wb = load_workbook(f, data_only=True)
        for name in wb.sheetnames:
            if '파크팩터' in name:
                year = int(name.split('_')[1])
                ws = wb[name]
                pf_all[year] = {}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[0] != '팀':
                        pf_all[year][row[0]] = float(row[-1])
        wb.close()
    pf_all[2025] = PF_2025
    return pf_all


def get_pf(pf_data, year, team):
    if year in pf_data and team in pf_data[year]:
        return pf_data[year][team]
    mapped = TEAM_NAME_MAP.get((year, team))
    if mapped and year in pf_data and mapped in pf_data[year]:
        return pf_data[year][mapped]
    return None


def parse_ip(ip_str):
    if ip_str is None:
        return 0.0
    ip_str = str(ip_str).strip()
    if not ip_str:
        return 0.0
    try:
        return float(ip_str)
    except ValueError:
        pass
    parts = ip_str.split()
    if len(parts) == 2:
        try:
            whole = int(parts[0])
            frac = parts[1].split('/')
            if len(frac) == 2:
                return whole + int(frac[0]) / int(frac[1])
        except (ValueError, ZeroDivisionError):
            pass
    return 0.0


def calculate_era_plus(db_path, pf_data, min_ip_ratio=1.0):
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    results = {}

    for year in range(1982, 2026):
        table = f'kbo_pitching_basic1_{year}'
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,))
        if not cursor.fetchone():
            continue

        cursor.execute(f"SELECT Player_ID, 선수명, 팀명, ERA, IP, ER, R, G, W, L, SV FROM [{table}]")
        rows = cursor.fetchall()
        if not rows:
            continue

        total_er, total_ip = 0, 0.0
        for row in rows:
            er = row[5] if row[5] is not None else 0
            total_er += er
            total_ip += parse_ip(row[4])

        if total_ip == 0:
            continue

        lg_era = (total_er * 9) / total_ip
        max_g = max((row[7] for row in rows if row[7]), default=144)
        min_ip = max_g * min_ip_ratio

        year_results = []
        for row in rows:
            pid, name, team, era, ip_str, er, r, g, w, l, sv = row
            ip = parse_ip(ip_str)
            if ip < min_ip or era is None or era <= 0:
                continue

            pf = get_pf(pf_data, year, team)
            if pf is None:
                continue

            era_plus = 100 * (lg_era / era) * (pf / 100)
            year_results.append({
                'Player_ID': pid, '선수명': name, '팀명': team,
                'ERA': era, 'IP': ip_str, 'IP_num': ip,
                'G': g, 'W': w, 'L': l,
                'lgERA': round(lg_era, 2), 'PF': pf, 'ERA+': round(era_plus, 1),
            })

        year_results.sort(key=lambda x: -x['ERA+'])
        results[year] = year_results

    conn.close()
    return results


def print_era_plus(results, year, top_n=20):
    if year not in results or not results[year]:
        print(f"{year}년 데이터 없음")
        return
    data = results[year]
    lg_era = data[0]['lgERA']
    print(f"\n{'='*70}")
    print(f"  {year} KBO ERA+ (리그 ERA: {lg_era})")
    print(f"{'='*70}")
    print(f"{'#':>3} {'선수명':>8} {'팀':>6} {'ERA':>6} {'IP':>8} {'W-L':>6} {'PF':>6} {'ERA+':>7}")
    print(f"{'-'*70}")
    for i, p in enumerate(data[:top_n], 1):
        wl = f"{p['W']}-{p['L']}"
        print(f"{i:>3} {p['선수명']:>8} {p['팀명']:>6} {p['ERA']:>6.2f} {p['IP']:>8} {wl:>6} {p['PF']:>6.1f} {p['ERA+']:>7.1f}")


if __name__ == "__main__":
    pf_files = [
        "kbo_statiz_data_1982_1998_.xlsx",
        "kbo_statiz_data_1999_2009_.xlsx",
        "kbo_statiz_data_2010_2021_.xlsx",
        "kbo_statiz_data_2022_2024_.xlsx",
    ]
    pf_data = load_park_factors(pf_files)
    print(f"파크팩터: {len(pf_data)}개 연도")

    results = calculate_era_plus("kbo_data.db", pf_data)
    print(f"ERA+ 계산: {len(results)}개 연도")

    for year in [2025, 2024, 2020, 2015, 2010, 2005, 2000, 1995, 1990, 1985]:
        if year in results:
            print_era_plus(results, year, top_n=10)
